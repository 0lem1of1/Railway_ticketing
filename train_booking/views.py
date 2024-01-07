from django.shortcuts import render,redirect,get_object_or_404
from .models import Train,Booking,Wallet,Journey,Passenger,Station
from user_app.models import User
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from .forms import JourneySearchForm,AddTrainForm,PassengerForm,PassengerFormSet,inlineformset_factory,AddMoneyForm,TrainUpdateForm
from django.http import HttpResponseRedirect,HttpResponse
from django.urls import reverse
from django.contrib import messages
from datetime import timezone,timedelta,datetime
import openpyxl
from django.core.mail import send_mail
from django.conf import settings

@login_required
def upcoming_journeys(request):
    user = request.user
    current_datetime = datetime.now()

    upcoming_bookings = Booking.objects.filter(
        user=user,
        journey__departure_date__gte=current_datetime,
        is_cancelled=False
    ).order_by('journey__departure_date')

    past_bookings = Booking.objects.filter(
        user=user,
        journey__departure_date__lt=current_datetime,
        is_cancelled=False
    ).order_by('-journey__departure_date')

    context = {
        'upcoming_bookings': upcoming_bookings,
        'past_bookings': past_bookings,
        'current_datetime': current_datetime,
    }

    return render(request, 'train_booking/upcoming_journeys.html', context)

@login_required
def cancel_booking(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id, user=request.user)

    if booking.is_cancelled:
        messages.error(request, "This booking has already been cancelled.")
    else:
        booking.is_cancelled = True
        booking.save()
        messages.success(request, "Booking cancelled successfully.")

    return redirect('home')

def edit_passenger(request, passenger_id):
    passenger = get_object_or_404(Passenger, id=passenger_id)

    if request.method == 'POST':
        form = PassengerForm(request.POST, instance=passenger)
        if form.is_valid():
            form.save()
            # Redirect to the passenger information page or any other appropriate view
            return redirect('home')
    else:
        form = PassengerForm(instance=passenger)

    return render(request, 'train_booking/edit_passenger.html', {'form': form})

def search_journey(request):
    form = JourneySearchForm(request.GET)

    if form.is_valid():
        source_station = form.cleaned_data['source_station']
        destination_station = form.cleaned_data['destination_station']
        departure_date = form.cleaned_data['departure_date']

        
        # Redirect to the search results page with query parameters
        url = reverse('search_results')
        url += f'?source_station={source_station}&destination_station={destination_station}&departure_date={departure_date}'
        return redirect(url)

    return render(request, 'train_booking/search_trains.html', {'form': form, 'errors': form.errors})


# search_results view
def search_results(request):
    # Retrieve the search criteria from query parameters
    source_station_name = request.GET.get('source_station')
    destination_station_name = request.GET.get('destination_station')
    departure_date = request.GET.get('departure_date')
    print(f"Source Station: {source_station_name}, Destination Station: {destination_station_name}, Departure Date: {departure_date}")

    # Find the Station objects corresponding to the provided names
    source_station = get_object_or_404(Station, name=source_station_name)
    destination_station = get_object_or_404(Station, name=destination_station_name)

    # Perform the search based on the query parameters
    journeys = Journey.objects.filter(
        train__source_station=source_station,
        train__route__in=[destination_station],
        departure_date=departure_date,
    )

    return render(request, 'train_booking/search_results.html', {'journeys': journeys})


@login_required
def create_booking(request):
    print(request.GET)
    journey_id = request.GET.get('journey_id')

    if request.method == 'POST':
        passenger_formset = PassengerFormSet(request.POST)

        if passenger_formset.is_valid():
            journey_id = request.POST.get('journey_id')
            # Handle the creation and initialization of the new booking
            print("POST data:", request.POST)

            journey = Journey.objects.get(pk=journey_id)
            booking = Booking.objects.create(user=request.user, journey=journey, amount_paid=0,seat_class=request.POST.get('seat_class'))

            # Iterate over the forms to handle each passenger
            for form in passenger_formset.forms:
                if form.is_valid():
                    Passenger.objects.create(
                        booking=booking,
                        name=form.cleaned_data['name'],
                        age=form.cleaned_data['age'],
                        gender=form.cleaned_data['gender'],
                    )

            booking.amount_paid = booking.journey.train.fare * passenger_formset.total_form_count()
            booking.save()

            wallet = Wallet.objects.get(user=request.user)
            wallet.subtract_money(booking.amount_paid)
            wallet.save()

            # Send email to the user
            send_booking_confirmation_email(request.user.email, booking)


            return redirect('success_page')

    else:
        passenger_formset = PassengerFormSet()

    return render(request, 'train_booking/create_booking.html', {'passenger_formset': passenger_formset, 'journey_id': journey_id})

DEFAULT_FROM_EMAIL = 'tijyp2+5kjb6usztxqw4@sharklasers.com'
def send_booking_confirmation_email(to_email, booking):
    try:
        subject = 'Booking Confirmation'
        message = f'Thank you for booking with us!\n\nBooking Details:\nTrain: {booking.journey}\nAmount Paid: {booking.amount_paid}\nSeat Class: {booking.seat_class}\n\nPassengers:\n'
        for passenger in booking.passenger_set.all():
            message += f'{passenger.name} - Age: {passenger.age} - Gender: {passenger.gender}\n'
        
        send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [to_email])
        print(f"Email sent to {to_email}")
        

    except Exception as e:
        print(f"Error sending email: {e}")
def success_page(request):
    passenger_count = request.GET.get('passenger_count', 0)
    return render(request, 'train_booking/success_page.html', {'passenger_count': passenger_count})

@login_required
def wallet_interface(request):
    
    wallet = Wallet.objects.get(user=request.user)

    if request.method == 'POST':
        form = AddMoneyForm(request.POST)
        if form.is_valid():
            amount = form.cleaned_data['amount']
            wallet.add_money(amount)
            wallet.save()
    else:
        form = AddMoneyForm()

    return render(request, 'train_booking/wallet_interface.html', {'wallet': wallet, 'form': form})

@login_required
def add_train(request):
    # Check if the logged-in user has the role ADMIN
    if request.user.role != 'ADMIN':
        return render(request, 'unauthorized_access.html')

    if request.method == 'POST':
        form = AddTrainForm(request.POST)
        if form.is_valid():

            # Process the route input and split it into a list
            route_input = form.cleaned_data['route_input']
            route_list = [station.strip() for station in route_input.split(',')]

            # Get or create Station objects for source and destination
            source_station,created_source = Station.objects.get_or_create(name=form.cleaned_data['source_station'])
            destination_station,created_destination = Station.objects.get_or_create(name=form.cleaned_data['destination_station'])


            # Create a Train object but don't save it yet
            new_train = Train(
                name=form.cleaned_data['name'],
                source_station=source_station,
                destination_station=destination_station,
                departure_time=form.cleaned_data['departure_time'],
                arrival_time=form.cleaned_data['arrival_time'],
                days_of_the_week=form.cleaned_data['days_of_the_week'],
                total_seats=form.cleaned_data['total_seats'],
                fare=form.cleaned_data['fare'],
            )
            
            # Save the Train object to the database
            new_train.save()

            # Process and save the route list
            new_train.set_route(route_list)

            # Create journeys for the next month
            new_train.create_journeys_for_next_months()

            return redirect('home_admin')
    else:
        form = AddTrainForm()

    return render(request, 'train_booking/add_train.html', {'form': form})



def bookings(request, train_id):
    train = get_object_or_404(Train, pk=train_id)
    bookings = Booking.objects.filter(journey__train=train, is_cancelled=False)
    return render(request, 'train_booking/bookings.html', {'train': train, 'bookings': bookings})


login_required
def train_list(request):
    trains = Train.objects.all()
    return render(request, 'train_booking/train_list.html', {'trains': trains})

@login_required
def update_train(request, train_id):
    train = get_object_or_404(Train, pk=train_id)

    if request.method == 'POST':
        form =TrainUpdateForm(request.POST, instance=train)
        if form.is_valid():
            form.save()
            return redirect('train_list')
    else:
        form = TrainUpdateForm(instance=train)

    return render(request, 'train_booking/update_train.html', {'form': form, 'train': train})


from django.db import transaction

@login_required
def delete_train(request, train_id):
    train = get_object_or_404(Train, pk=train_id)

    if request.method == 'POST':
        if request.POST.get('confirm_delete'):
            with transaction.atomic():
                # Step 1: Identify bookings for the train
                bookings = Booking.objects.filter(journey__train=train, is_cancelled=False)

                # Step 2: Optionally, mark the journeys as canceled
                # If Journey model has an 'is_cancelled' field, set it to True
                # This check is added to prevent unnecessary processing if already canceled
                for journey in train.journey_set.all():
                    if hasattr(journey, 'is_cancelled') and not journey.is_cancelled:
                        journey.is_cancelled = True
                        journey.save()

                # Step 3: Cancel bookings and refund users
                for booking in bookings:
                    user = booking.user
                    wallet = Wallet.objects.get(user=user)

                    # Refund the amount paid for the booking
                    wallet.add_money(booking.amount_paid)
                    wallet.save()

                    # Cancel the booking
                    booking.is_cancelled = True
                    booking.save()

                # Finally, delete the train
                train.delete()

                messages.success(request, 'Train deleted successfully. Refunds processed.')
            return redirect('train_list')
        else:
            messages.warning(request, 'Deletion canceled. No changes were made.')
            return redirect('train_list')

    return render(request, 'train_booking/delete_train.html', {'train': train})

def export_bookings_excel(request,train_id):
    train = get_object_or_404(Train,pk=train_id)

    # Fetch your booking data (replace this with your actual data retrieval logic)
    bookings = Booking.objects.filter(journey__train=train,is_cancelled=False)

    # Create an Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Add headers to the worksheet
    headers = ["User", "Train", "Amount Paid", "Seat Class", "Passenger Name", "Gender", "Age"]
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Add booking and passenger data to the worksheet
    for row_num, booking in enumerate(bookings, 2):
        worksheet.cell(row=row_num, column=1, value=booking.user.username)
        worksheet.cell(row=row_num, column=2, value=str(booking.journey))
        worksheet.cell(row=row_num, column=3, value=booking.amount_paid)
        worksheet.cell(row=row_num, column=4, value=booking.seat_class)

        # Iterate over passengers for the current booking
        for passenger in booking.passenger_set.all():
            worksheet.cell(row=row_num, column=5, value=passenger.name)
            worksheet.cell(row=row_num, column=6, value=passenger.gender)
            worksheet.cell(row=row_num, column=7, value=passenger.age)

            # Increment row number for the next passenger (if any)
            row_num += 1

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=bookings.xlsx'
    workbook.save(response)

    return response

